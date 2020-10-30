#! usr/bin/python
# -*- coding:utf-8 -*-
"""
掷骰子 v1.1
1、欢迎进入游戏
2、输入用户名密码进入游戏，默认是没有币的
3、提示用户充值买币（100块钱3个币，充值必须是100的倍数，充值不成功可以再次充值）
4、玩一局需要扣除2个币，猜大小（系统产生随机数）
5、只要猜对了奖励1个币，可以继续玩，也可以主动退出，或者没有币强制退出
"""


# 主函数,负责调用函数执行功能
def main():
    print('*' * 5 + "欢迎进入游戏" + "*" * 5)
    currency1 = 0
    status = True
    while True:
        if status:
            user1 = input('注册/登陆/退出(1/2/0):')
            if user1 == "1":
                bol = registered('../test/server.xlsx', 'Sheet1')
                if bol[1] != None:
                    currency1 = bol[1]
                status = False
            elif user1 == "2":
                bol = authentication()
                if bol[1] != None:
                    currency1 = bol[1]
                status = False
            elif user1 == "0":
                return None
        else:
            if bol[0]:
                if currency1 > 2:
                    game(currency1)
                    carry_on = input('是否继续游戏(y/n)')
                    if carry_on == 'y':
                        pass
                    else:
                        print('再见！')
                        return None
                else:
                    currency1 = recharge()


def public(url, Sheet):
    lis = []
    from openpyxl import load_workbook
    server = load_workbook(url)
    sheet = server[Sheet]
    bol = True
    i = 1
    cell1 = 'A'
    cell2 = 'B'
    cell3 = 'C'
    while bol:
        information = {}
        none = None
        cell1 = cell1 + str(i)
        cell2 = cell2 + str(i)
        cell3 = cell3 + str(i)
        if sheet[cell1].value == none:
            server.close()
            return lis, i
        else:
            information['name'] = sheet[cell1].value
            information['password'] = sheet[cell2].value
            information['rechrages'] = sheet[cell3].value
            lis.append(information)
            information = {}
        i += 1


# 身份验证函数，主要负责用户的身份验证，读取用户的用户名，游戏币等信息
def authentication():
    """
    登录函数负责严重用户的登录信息
    :return:
    """
    group = public('../test/server.xlsx', 'Sheet1')
    verification = group[0]
    name = input('请输入用户名:')
    password = input('请输入密码：')
    for dictionary in verification:
        list1 = [x for x in dictionary.values()]
        if list1[0] == name and list1[1] == password:
            print('登录成功！')
            return True, list1[2]
        else:
            print('登录失败！')
            return False, None


# 注册函数
def registered(url, sheet):
    """
    注册函数，负责用户的注册信息，吧注册的信息写入到表格
    :return: None
    """
    import re
    from openpyxl import load_workbook
    server = load_workbook(url)
    sheet1 = server[sheet]
    sum2 = str(public('../test/server.xlsx', 'Sheet1')[1])
    letters = ['A', 'B', 'C']
    letters = [x + sum2 for x in letters]
    name = input('请输入用户名(4-6个字符)：')
    password = input('请输入密码：')
    result = re.search(r'.{4,6}', name)
    sheet1[letters[0]] = name
    sheet1[letters[1]] = password
    print('注册成功')
    server.save(url)
    server.close()
    return True, None


# 充值函数，负责用户的充值
def recharge():
    """
    充值函数，负责用户的充值，充值只能是100的倍数
    :return: 充值的币的个数
    """
    money = int(input('请输入你要充值的钱数(必须是100的倍数)：'))
    if money % 100 != 0:
        print('请输入100的倍数！')
    else:
        currency = (money / 100) * 3
        print('充值成功%d个游戏币!' % currency)
        return currency


# 游戏函数，负责游戏的进行，当游戏币不足是退出，用户主动退出
def game(time):
    """
    游戏竞猜函数，生成一个1-6的随机数竞猜正确加3个币错误扣三个
    :param time:
    :return:
    """
    import random
    if time > 2:
        print('游戏开始，猜大小！')
        time -= 2
        mum = float(input('请输入你要竞猜的数字(必须是整数1-6):'))
        sums = random.randint(1, 6)
        if mum == sums:
            print('结果为：%d\n竞猜正确恭喜您，获得三个币的奖励！' % sums)
            time += 5
        else:
            print('结果为：%d\n再接再厉，下一把说不定有好运!' % sums)
        return time
    else:
        print('游戏币不足，请充值！')
        return time


# 程序入口
if __name__ == '__main__':
    main()
