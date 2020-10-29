#! usr/bin/python
# -*- coding:utf-8 -*-
"""
创建一个工作薄：wb = openpyxl.Workbook()
新增一个sheet表单：wb.create_sheet('test_case')
保存case.xlsx文件：wb.save('cases.xlsx')
打开工作簿：wb = openpyxl.load_workbook('cases.xlsx')
选取表单：sh = wb['Sheet1'
读取第一行、第一列的数据：ce = sh.cell(row = 1,column = 1)
按行读取数据：row_data = list(sh.rows)
关闭工作薄：wb.close()
按列读取数据：columns_data = list(sh.columns)
写入数据之前，该文件一定要处于关闭状态
写入第一行、第四列的数据 value = 'result'：sh.cell(row = 1,column = 4,value = 'result')
获取最大行总数、最大列总数：sh.max_row、sh.max_column
del 删除表单的用法：del wb['sheet_name']
remove 删除表单的用法：sh = wb['sheet_name'] wb.remove(sh)
"""
import base64
from openpyxl import load_workbook

# 打开文件，参数是路径
wb = load_workbook('./server.xlsx')
# 获取文件的页面
ws = wb['Sheet1']
# a = 'A'
# s = 0
# for sheet in wb:
#     sa = a + str(s)
#     s += 1
#     # 获取指定行
#     print(sheet[1])
#     # 获取指定列
#     print(sheet['A'])
#     # 获取特定单元格
#     print(sheet['A1'])

# while s < 3:
#     sa = a + str(s)
#     # 获取指定列
#     print(sheet[sa].value)
#     s += 1
# # 关闭文件
# wb.close()
#
# lis = ['A', 'B', 'C', 'D']
#
# for grid in lis:
#     for one in range(30):
#         cell = grid + str(one)
#         print(cell)

ws['A1'] = "测试数据"
wb.save('./server.xlsx')
